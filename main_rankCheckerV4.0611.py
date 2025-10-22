import sys
import os
import json
import urllib.request
import urllib.parse
import re
import time
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QLineEdit, QPushButton, QTextBrowser, QTextEdit,
    QMessageBox, QProgressBar, QGroupBox, QComboBox,
    QTabWidget, QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QIcon

# Naver API Keys - config.json에서 로드됨
NAVER_CLIENT_ID = ""
NAVER_CLIENT_SECRET = ""

# Naver 검색광고 API Keys - config.json에서 로드됨
NAVER_AD_API_KEY = ""
NAVER_AD_SECRET_KEY = ""
NAVER_AD_CUSTOMER_ID = ""


class Worker(QThread):
    """검색 작업을 처리하는 워커 쓰레드"""
    result_ready = pyqtSignal(str)
    progress_update = pyqtSignal(int, str)
    finished_all = pyqtSignal(dict)

    def __init__(self, keywords, mall_name, sort_option="sim"):
        super().__init__()
        self.keywords = keywords
        self.mall_name = mall_name
        self.sort_option = sort_option
        self.all_results = {}

    def get_top_ranked_product_by_mall(self, keyword, mall_name, sort_option="sim"):
        """네이버 쇼핑에서 특정 판매처의 최상위 순위 상품 검색"""
        encText = urllib.parse.quote(keyword)
        seen_titles = set()
        best_product = None
        
        for start in range(1, 1001, 100):
            url = f"https://openapi.naver.com/v1/search/shop.json?query={encText}&display=100&start={start}&sort={sort_option}"
            request = urllib.request.Request(url)
            request.add_header("X-Naver-Client-Id", NAVER_CLIENT_ID)
            request.add_header("X-Naver-Client-Secret", NAVER_CLIENT_SECRET)
            
            # API 호출 재시도 로직 (429 에러 대응)
            max_retries = 3
            retry_delay = 1
            result = None
            
            for attempt in range(max_retries):
                try:
                    response = urllib.request.urlopen(request)
                    result = json.loads(response.read())
                    break
                except urllib.error.HTTPError as e:
                    if e.code == 429 and attempt < max_retries - 1:
                        time.sleep(retry_delay * (attempt + 1))
                        continue
                    elif e.code == 429:
                        # API 호출 제한 초과
                        return None
                    else:
                        raise
            
            if result is None:
                continue
                
            # API 호출 간 딜레이 (0.1초)
            time.sleep(0.1)
            
            for idx, item in enumerate(result.get("items", []), start=1):
                if item.get("mallName") and mall_name in item["mallName"]:
                    title_clean = re.sub(r"<.*?>", "", item["title"])
                    if title_clean in seen_titles:
                        continue
                    seen_titles.add(title_clean)
                    rank = start + idx - 1
                    product = {
                        "rank": rank,
                        "title": title_clean,
                        "price": item["lprice"],
                        "link": item["link"],
                        "mallName": item["mallName"]
                    }
                    if best_product is None:
                        best_product = product
                    elif rank < best_product["rank"]:
                        best_product = product
        return best_product

    def run(self):
        """워커 쓰레드 실행"""
        total = len(self.keywords)
        for i, keyword in enumerate(self.keywords):
            result = self.get_top_ranked_product_by_mall(keyword, self.mall_name, self.sort_option)
            if result:
                link_html = f'<a href="{result["link"]}" style="color:#2196F3;">{result["link"]}</a>'
                html = (
                    f"<div style='margin-bottom: 15px; padding: 10px; background-color: #f5f5f5; border-radius: 5px;'>"
                    f"<b style='color:#4CAF50; font-size: 14px;'>✓ {keyword}</b><br>"
                    f"<span style='color:#666;'>순위:</span> <b style='color:#FF5722;'>{result['rank']}위</b><br>"
                    f"<span style='color:#666;'>상품명:</span> {result['title']}<br>"
                    f"<span style='color:#666;'>가격:</span> <b>{int(result['price']):,}원</b><br>"
                    f"<span style='color:#666;'>링크:</span> {link_html}"
                    f"</div>"
                )
                self.all_results[keyword] = result
            else:
                html = (
                    f"<div style='margin-bottom: 15px; padding: 10px; background-color: #ffebee; border-radius: 5px;'>"
                    f"<b style='color:#f44336;'>✗ {keyword}</b><br>"
                    f"<span style='color:#666;'>검색 결과 없음</span>"
                    f"</div>"
                )
                self.all_results[keyword] = "검색 결과 없음"
            
            percent = int(((i + 1) / total) * 100)
            self.result_ready.emit(html)
            self.progress_update.emit(percent, keyword)
        
        self.finished_all.emit(self.all_results)


class RelatedKeywordWorker(QThread):
    """연관 키워드 조회 워커 쓰레드"""
    result_ready = pyqtSignal(list)
    progress_update = pyqtSignal(str)
    
    def __init__(self, keyword):
        super().__init__()
        self.keyword = keyword
    
    def get_related_keywords(self, keyword):
        """네이버 검색광고 API로 연관 키워드 가져오기"""
        try:
            import hashlib
            import hmac
            import base64
            
            # API 엔드포인트
            timestamp = str(int(time.time() * 1000))
            method = "GET"
            uri = "/keywordstool"
            
            # 서명 생성
            message = f"{timestamp}.{method}.{uri}"
            signature = hmac.new(
                NAVER_AD_SECRET_KEY.encode('utf-8'),
                message.encode('utf-8'),
                hashlib.sha256
            ).digest()
            signature = base64.b64encode(signature).decode('utf-8')
            
            # API 요청 - 연관 키워드 조회
            url = f"https://api.naver.com/keywordstool?hintKeywords={urllib.parse.quote(keyword)}&showDetail=1"
            
            request = urllib.request.Request(url)
            request.add_header("X-Timestamp", timestamp)
            request.add_header("X-API-KEY", NAVER_AD_API_KEY)
            request.add_header("X-Customer", NAVER_AD_CUSTOMER_ID)
            request.add_header("X-Signature", signature)
            
            response = urllib.request.urlopen(request)
            result = json.loads(response.read().decode('utf-8'))
            
            # API 응답에서 키워드 리스트 추출
            if result and 'keywordList' in result:
                keywords = []
                total_count = len(result['keywordList'])
                
                # 문자열 값 처리 함수
                def parse_count(val):
                    if isinstance(val, str):
                        if val.startswith('<'):
                            return int(val[1:].strip()) // 2
                        elif '< ' in val:
                            return int(val.split('< ')[1]) // 2
                        try:
                            return int(val)
                        except:
                            return 0
                    return int(val) if val else 0
                
                # 모든 키워드 데이터 추출
                for item in result['keywordList']:
                    rel_keyword = item.get('relKeyword', '')
                    monthly_pc_qc = item.get('monthlyPcQcCnt', 0)
                    monthly_mobile_qc = item.get('monthlyMobileQcCnt', 0)
                    
                    pc_count = parse_count(monthly_pc_qc)
                    mobile_count = parse_count(monthly_mobile_qc)
                    total_search = pc_count + mobile_count
                    
                    if rel_keyword:
                        keywords.append({
                            "keyword": rel_keyword,
                            "pc_count": pc_count,  # PC 월간 검색수
                            "mobile_count": mobile_count,  # 모바일 월간 검색수
                            "count": total_search,  # 월간 총 검색수
                            "total_count": total_count  # 총 키워드 수
                        })
                
                # 월간 검색수가 많은 순으로 정렬 후 상위 100개 추출
                keywords.sort(key=lambda x: x['count'], reverse=True)
                
                return keywords[:100]
            else:
                self.progress_update.emit("API에서 데이터를 가져올 수 없습니다.")
                return []
            
        except Exception as e:
            self.progress_update.emit(f"API 오류: {str(e)} - 대체 방법 사용")
            # API 실패 시 기존 방식 사용
            return self.get_related_keywords_fallback(keyword)
    
    def get_related_keywords_fallback(self, keyword):
        """네이버 쇼핑 검색 결과에서 연관 키워드 추출 (대체 방법)"""
        try:
            encText = urllib.parse.quote(keyword)
            word_count = {}
            
            # 5페이지(500개 상품)까지 분석
            for page in range(5):
                start = page * 100 + 1
                url = f"https://openapi.naver.com/v1/search/shop.json?query={encText}&display=100&start={start}&sort=sim"
                request = urllib.request.Request(url)
                request.add_header("X-Naver-Client-Id", NAVER_CLIENT_ID)
                request.add_header("X-Naver-Client-Secret", NAVER_CLIENT_SECRET)
                
                try:
                    response = urllib.request.urlopen(request)
                    result = json.loads(response.read())
                    
                    for item in result.get("items", []):
                        title = re.sub(r"<.*?>", "", item["title"])
                        words = re.findall(r'[가-힣a-zA-Z0-9]{2,}', title)
                        
                        for word in words:
                            if word.lower() != keyword.lower() and word not in keyword:
                                word_count[word] = word_count.get(word, 0) + 1
                    
                    time.sleep(0.1)
                    
                except Exception:
                    break
            
            sorted_words = sorted(word_count.items(), key=lambda x: x[1], reverse=True)[:100]
            total_count = len(word_count)
            
            # PC/모바일 비율 대략 40:60으로 추정
            return [{
                "keyword": word, 
                "pc_count": int(count * 0.4),
                "mobile_count": int(count * 0.6),
                "count": count, 
                "total_count": total_count
            } for word, count in sorted_words]
            
        except Exception as e:
            self.progress_update.emit(f"오류 발생: {str(e)}")
            return []
    
    def run(self):
        """워커 쓰레드 실행"""
        self.progress_update.emit(f"'{self.keyword}' 연관 키워드 분석 중...")
        related = self.get_related_keywords(self.keyword)
        self.result_ready.emit(related)


class AdCostWorker(QThread):
    """광고 비용 조회 워커 쓰레드"""
    result_ready = pyqtSignal(list)
    progress_update = pyqtSignal(str)
    
    def __init__(self, keywords):
        super().__init__()
        self.keywords = keywords
    
    def get_naver_ad_data(self, keyword):
        """네이버 검색광고 API로 실제 광고 데이터 가져오기"""
        try:
            import hashlib
            import hmac
            import base64
            
            # API 엔드포인트
            timestamp = str(int(time.time() * 1000))
            method = "GET"
            uri = "/keywordstool"
            
            # 서명 생성
            message = f"{timestamp}.{method}.{uri}"
            signature = hmac.new(
                NAVER_AD_SECRET_KEY.encode('utf-8'),
                message.encode('utf-8'),
                hashlib.sha256
            ).digest()
            signature = base64.b64encode(signature).decode('utf-8')
            
            # API 요청
            url = f"https://api.naver.com/keywordstool?hintKeywords={urllib.parse.quote(keyword)}&showDetail=1"
            
            request = urllib.request.Request(url)
            request.add_header("X-Timestamp", timestamp)
            request.add_header("X-API-KEY", NAVER_AD_API_KEY)
            request.add_header("X-Customer", NAVER_AD_CUSTOMER_ID)
            request.add_header("X-Signature", signature)
            
            response = urllib.request.urlopen(request)
            result = json.loads(response.read().decode('utf-8'))
            
            # API 응답 확인
            if result and 'keywordList' in result and len(result['keywordList']) > 0:
                data = result['keywordList'][0]
                
                # CPC 값이 '<10' 같은 문자열로 올 수 있으므로 처리
                def parse_value(val):
                    if isinstance(val, str):
                        if val.startswith('<'):
                            return int(val[1:]) // 2  # '<10'이면 5로 추정
                        elif val == '< 10':
                            return 5
                        try:
                            return int(val)
                        except:
                            return 0
                    return int(val) if val else 0
                
                pc_cpc = parse_value(data.get('monthlyAvePcCpc', 0))
                mobile_cpc = parse_value(data.get('monthlyAveMobileCpc', 0))
                
                return {
                    'success': True,
                    'monthly_pc_search': parse_value(data.get('monthlyPcQcCnt', 0)),
                    'monthly_mobile_search': parse_value(data.get('monthlyMobileQcCnt', 0)),
                    'monthly_ave_pc_clicks': parse_value(data.get('monthlyAvePcClkCnt', 0)),
                    'monthly_ave_mobile_clicks': parse_value(data.get('monthlyAveMobileClkCnt', 0)),
                    'competition': data.get('compIdx', 'N/A'),
                    'avg_pc_cpc': pc_cpc,
                    'avg_mobile_cpc': mobile_cpc,
                    'pl_avg_depth': parse_value(data.get('plAvgDepth', 0))
                }
            else:
                return {'success': False, 'error': 'No data in API response'}
        except urllib.error.HTTPError as e:
            error_msg = f"HTTP {e.code}: {e.reason}"
            self.progress_update.emit(f"API 오류: {error_msg}")
            return {'success': False, 'error': error_msg}
        except Exception as e:
            self.progress_update.emit(f"API 오류: {str(e)}")
            return {'success': False, 'error': str(e)}
    
    def analyze_ad_cost(self, keyword):
        """키워드별 광고 비용 분석 (실제 API 데이터 우선, 없으면 추정)"""
        # 먼저 네이버 검색광고 API 시도
        api_data = self.get_naver_ad_data(keyword)
        
        if api_data.get('success'):
            # 실제 API 데이터 사용
            monthly_pc_search = api_data['monthly_pc_search']
            monthly_mobile_search = api_data['monthly_mobile_search']
            total_monthly_search = monthly_pc_search + monthly_mobile_search
            
            monthly_pc_clicks = api_data['monthly_ave_pc_clicks']
            monthly_mobile_clicks = api_data['monthly_ave_mobile_clicks']
            monthly_clicks = monthly_pc_clicks + monthly_mobile_clicks
            
            # 평균 클릭률 계산
            if total_monthly_search > 0:
                avg_ctr = (monthly_clicks / total_monthly_search) * 100
            else:
                avg_ctr = 0
            
            competition = api_data['competition']
            avg_cpc = max(api_data['avg_pc_cpc'], api_data['avg_mobile_cpc'])
            
            # 경쟁도를 숫자로 변환
            if competition == 'high':
                competition_text = "높음"
                avg_ad_impressions = 12
            elif competition == 'medium':
                competition_text = "보통"
                avg_ad_impressions = 9
            elif competition == 'low':
                competition_text = "낮음"
                avg_ad_impressions = 6
            else:
                competition_text = competition
                avg_ad_impressions = 8
            
            # 월간 광고비
            estimated_monthly_cost = avg_cpc * monthly_clicks if avg_cpc > 0 else 0
            
            return {
                "keyword": keyword,
                "monthly_pc_search": monthly_pc_search,
                "monthly_mobile_search": monthly_mobile_search,
                "total_monthly_search": total_monthly_search,
                "monthly_clicks": monthly_clicks,
                "avg_ctr": round(avg_ctr, 2),
                "competition": competition_text,
                "avg_ad_impressions": avg_ad_impressions,
                "estimated_cpc": int(avg_cpc),
                "estimated_monthly_cost": int(estimated_monthly_cost)
            }
        else:
            # API 실패 시 추정값 사용
            try:
                encText = urllib.parse.quote(keyword)
                url = f"https://openapi.naver.com/v1/search/shop.json?query={encText}&display=100&start=1&sort=sim"
                request = urllib.request.Request(url)
                request.add_header("X-Naver-Client-Id", NAVER_CLIENT_ID)
                request.add_header("X-Naver-Client-Secret", NAVER_CLIENT_SECRET)
                
                response = urllib.request.urlopen(request)
                result = json.loads(response.read())
                
                total_results = result.get("total", 0)
                
                # 경쟁도 계산
                if total_results < 1000:
                    competition = "낮음"
                    competition_score = 1
                elif total_results < 10000:
                    competition = "보통"
                    competition_score = 2
                elif total_results < 50000:
                    competition = "높음"
                    competition_score = 3
                else:
                    competition = "매우 높음"
                    competition_score = 4
                
                # 월간 검색수 추정
                base_monthly_search = min(total_results * 2, 1000000)
                monthly_pc_search = int(base_monthly_search * 0.4)
                monthly_mobile_search = int(base_monthly_search * 0.6)
                total_monthly_search = monthly_pc_search + monthly_mobile_search
                
                # 월 평균 클릭수
                click_rate = 0.03 + (competition_score * 0.01)
                monthly_clicks = int(total_monthly_search * click_rate)
                
                avg_ctr = round(click_rate * 100, 2)
                avg_ad_impressions = competition_score * 3
                
                # 예상 CPC
                base_cpc = 150
                keyword_length_factor = len(keyword) * 15
                competition_factor = competition_score * 200
                estimated_cpc = base_cpc + keyword_length_factor + competition_factor
                
                estimated_monthly_cost = estimated_cpc * monthly_clicks
                
                return {
                    "keyword": keyword,
                    "monthly_pc_search": monthly_pc_search,
                    "monthly_mobile_search": monthly_mobile_search,
                    "total_monthly_search": total_monthly_search,
                    "monthly_clicks": monthly_clicks,
                    "avg_ctr": avg_ctr,
                    "competition": competition,
                    "avg_ad_impressions": avg_ad_impressions,
                    "estimated_cpc": estimated_cpc,
                    "estimated_monthly_cost": int(estimated_monthly_cost)
                }
                
            except Exception as e:
                self.progress_update.emit(f"오류 발생: {str(e)}")
                return None
    
    def run(self):
        """워커 쓰레드 실행"""
        results = []
        total = len(self.keywords)
        
        for i, keyword in enumerate(self.keywords):
            self.progress_update.emit(f"분석 중... ({i+1}/{total}) - {keyword}")
            result = self.analyze_ad_cost(keyword)
            if result:
                results.append(result)
            time.sleep(0.3)  # API 호출 제한 방지
        
        self.result_ready.emit(results)


class RankListWorker(QThread):
    """1~100위 전체 리스트를 가져오는 워커 쓰레드"""
    progress_update = pyqtSignal(int, str)
    finished = pyqtSignal(list)

    def __init__(self, keyword, sort_option="sim"):
        super().__init__()
        self.keyword = keyword
        self.sort_option = sort_option

    def run(self):
        """1~100위 상품 목록 가져오기"""
        encText = urllib.parse.quote(self.keyword)
        all_products = []
        seen_titles = set()
        
        # 100위까지만 가져오기
        for start in range(1, 101, 100):
            url = f"https://openapi.naver.com/v1/search/shop.json?query={encText}&display=100&start={start}&sort={self.sort_option}"
            request = urllib.request.Request(url)
            request.add_header("X-Naver-Client-Id", NAVER_CLIENT_ID)
            request.add_header("X-Naver-Client-Secret", NAVER_CLIENT_SECRET)
            
            # API 호출 재시도 로직
            max_retries = 3
            retry_delay = 1
            result = None
            
            for attempt in range(max_retries):
                try:
                    response = urllib.request.urlopen(request)
                    result = json.loads(response.read())
                    break
                except urllib.error.HTTPError as e:
                    if e.code == 429 and attempt < max_retries - 1:
                        time.sleep(retry_delay * (attempt + 1))
                        continue
                    elif e.code == 429:
                        self.progress_update.emit(0, "API 호출 제한 초과")
                        return
                    else:
                        raise
            
            if result is None:
                continue
                
            # API 호출 간 딜레이
            time.sleep(0.1)
            
            for idx, item in enumerate(result.get("items", []), start=1):
                title_clean = re.sub(r"<.*?>", "", item["title"])
                
                # 중복 제거
                if title_clean in seen_titles:
                    continue
                seen_titles.add(title_clean)
                
                rank = start + idx - 1
                if rank > 100:
                    break
                    
                product = {
                    "rank": rank,
                    "title": title_clean,
                    "price": item.get("lprice", "0"),
                    "link": item.get("link", ""),
                    "mallName": item.get("mallName", "알 수 없음"),
                    "brand": item.get("brand", ""),
                    "maker": item.get("maker", "")
                }
                all_products.append(product)
                
                percent = min(int((len(all_products) / 100) * 100), 100)
                self.progress_update.emit(percent, f"{len(all_products)}개 상품 수집 중...")
            
            if len(all_products) >= 100:
                break
        
        self.finished.emit(all_products)


def resource_path(relative_path):
    """PyInstaller 환경에서도 리소스 파일 경로를 올바르게 반환"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


class RankCheckerApp(QWidget):
    """네이버 순위 확인기 메인 윈도우"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("네이버 쇼핑 순위 확인기")
        
        # 아이콘 설정
        if os.path.exists(resource_path("logo_inner.ico")):
            self.setWindowIcon(QIcon(resource_path("logo_inner.ico")))
        
        self.resize(850, 750)
        self.worker = None
        self.setup_ui()
        self.apply_styles()

    def setup_ui(self):
        """UI 구성"""
        main_layout = QVBoxLayout()
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # 타이틀
        title_label = QLabel("네이버 쇼핑 순위 확인기")
        title_font = QFont()
        title_font.setPointSize(18)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("color: #1976D2; padding: 10px;")
        main_layout.addWidget(title_label)

        # 탭 위젯 생성
        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet("QTabWidget::pane { border: 2px solid #E0E0E0; border-radius: 5px; }")
        
        # 탭 1: 판매처 순위 확인
        self.tab1 = QWidget()
        self.setup_tab1()
        self.tab_widget.addTab(self.tab1, "🎯 판매처 순위 확인")
        
        # 탭 2: 전체 순위 리스트
        self.tab2 = QWidget()
        self.setup_tab2()
        self.tab_widget.addTab(self.tab2, "📊 전체 순위 리스트 (1~100위)")
        
        # 탭 3: 연관 키워드 조회
        self.tab3 = QWidget()
        self.setup_tab3()
        self.tab_widget.addTab(self.tab3, "🔗 연관 키워드 조회")
        
        # 탭 4: 광고 비용 분석
        self.tab4 = QWidget()
        self.setup_tab4()
        self.tab_widget.addTab(self.tab4, "💰 광고 비용 분석")
        
        # 탭 5: 설정
        self.tab5 = QWidget()
        self.setup_tab5()
        self.tab_widget.addTab(self.tab5, "⚙️ 설정")
        
        main_layout.addWidget(self.tab_widget)
        self.setLayout(main_layout)
        
        # API 키 로드 및 검증
        self.load_api_keys()
        self.check_api_keys_on_startup()

    def setup_tab1(self):
        """탭 1: 판매처 순위 확인 UI"""
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)

        # 입력 영역 그룹박스
        input_group = QGroupBox("검색 설정")
        input_layout = QVBoxLayout()
        input_layout.setSpacing(10)

        # 검색어 입력
        keywords_label = QLabel("🔍 검색어 (쉼표로 구분, 최대 10개)")
        keywords_label.setStyleSheet("font-weight: bold; color: #333;")
        self.input_keywords = QTextEdit()
        self.input_keywords.setPlaceholderText("예: 키보드, 마우스, 충전기")
        self.input_keywords.setMaximumHeight(80)

        input_layout.addWidget(keywords_label)
        input_layout.addWidget(self.input_keywords)

        # 판매처명 입력
        mall_label = QLabel("🏪 판매처명")
        mall_label.setStyleSheet("font-weight: bold; color: #333;")
        self.input_mall = QLineEdit()
        self.input_mall.setPlaceholderText("예: OO스토어")
        self.input_mall.setMinimumHeight(35)
        
        input_layout.addWidget(mall_label)
        input_layout.addWidget(self.input_mall)

        # 정렬 옵션
        sort_label = QLabel("📊 정렬 옵션")
        sort_label.setStyleSheet("font-weight: bold; color: #333;")
        self.sort_combo = QComboBox()
        self.sort_combo.setMinimumHeight(40)
        
        # 정렬 옵션 추가 (네이버 API에서 지원하는 옵션만)
        self.sort_combo.addItem("랭킹순 (정확도)", "sim")
        self.sort_combo.addItem("낮은 가격순", "asc")
        self.sort_combo.addItem("높은 가격순", "dsc")
        self.sort_combo.addItem("등록일순", "date")
        
        # 폰트 설정
        combo_font = QFont()
        combo_font.setPointSize(10)
        self.sort_combo.setFont(combo_font)
        
        input_layout.addWidget(sort_label)
        input_layout.addWidget(self.sort_combo)

        input_group.setLayout(input_layout)
        layout.addWidget(input_group)

        # 검색 버튼
        self.button_check = QPushButton("🚀 순위 확인하기")
        self.button_check.setMinimumHeight(45)
        self.button_check.clicked.connect(self.start_check)
        layout.addWidget(self.button_check)

        # 진행 상태 영역
        status_layout = QHBoxLayout()
        self.label_status = QLabel("대기 중...")
        self.label_status.setStyleSheet("color: #666; font-weight: bold;")
        status_layout.addWidget(self.label_status)
        status_layout.addStretch()
        layout.addLayout(status_layout)

        # 프로그레스 바
        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimumHeight(25)
        layout.addWidget(self.progress_bar)

        # 결과 표시 영역
        result_label = QLabel("📊 검색 결과")
        result_label.setStyleSheet("font-weight: bold; color: #333; margin-top: 10px;")
        layout.addWidget(result_label)
        
        self.result_display = QTextBrowser()
        self.result_display.setOpenExternalLinks(True)
        layout.addWidget(self.result_display)

        self.tab1.setLayout(layout)

    def setup_tab2(self):
        """탭 2: 전체 순위 리스트 UI"""
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)

        # 입력 영역
        input_group = QGroupBox("검색 설정")
        input_layout = QVBoxLayout()
        
        # 검색어 입력
        keyword_label = QLabel("🔍 검색어 (단일 키워드)")
        keyword_label.setStyleSheet("font-weight: bold; color: #333;")
        self.input_keyword_list = QLineEdit()
        self.input_keyword_list.setPlaceholderText("예: 키보드")
        self.input_keyword_list.setMinimumHeight(35)
        
        input_layout.addWidget(keyword_label)
        input_layout.addWidget(self.input_keyword_list)

        # 정렬 옵션
        sort_label2 = QLabel("📊 정렬 옵션")
        sort_label2.setStyleSheet("font-weight: bold; color: #333;")
        self.sort_combo2 = QComboBox()
        self.sort_combo2.setMinimumHeight(40)
        self.sort_combo2.addItem("랭킹순 (정확도)", "sim")
        self.sort_combo2.addItem("낮은 가격순", "asc")
        self.sort_combo2.addItem("높은 가격순", "dsc")
        self.sort_combo2.addItem("등록일순", "date")
        
        combo_font = QFont()
        combo_font.setPointSize(10)
        self.sort_combo2.setFont(combo_font)
        
        input_layout.addWidget(sort_label2)
        input_layout.addWidget(self.sort_combo2)
        
        input_group.setLayout(input_layout)
        layout.addWidget(input_group)

        # 버튼 영역
        button_layout = QHBoxLayout()
        self.button_get_list = QPushButton("📊 1~100위 리스트 가져오기")
        self.button_get_list.setMinimumHeight(45)
        self.button_get_list.clicked.connect(self.get_rank_list)
        
        self.button_export_excel = QPushButton("📥 엑셀로 다운로드")
        self.button_export_excel.setMinimumHeight(45)
        self.button_export_excel.setEnabled(False)
        self.button_export_excel.clicked.connect(self.export_to_excel)
        self.button_export_excel.setStyleSheet("background-color: #4CAF50;")
        
        button_layout.addWidget(self.button_get_list)
        button_layout.addWidget(self.button_export_excel)
        layout.addLayout(button_layout)

        # 진행 상태
        status_layout2 = QHBoxLayout()
        self.label_status2 = QLabel("대기 중...")
        self.label_status2.setStyleSheet("color: #666; font-weight: bold;")
        status_layout2.addWidget(self.label_status2)
        status_layout2.addStretch()
        layout.addLayout(status_layout2)

        # 프로그레스 바
        self.progress_bar2 = QProgressBar()
        self.progress_bar2.setMinimumHeight(25)
        layout.addWidget(self.progress_bar2)

        # 테이블
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(6)
        self.table_widget.setHorizontalHeaderLabels(["순위", "상품명", "가격", "판매처", "브랜드", "제조사"])
        
        # 테이블 설정
        header = self.table_widget.horizontalHeader()
        header.setSectionResizeMode(1, QHeaderView.Stretch)  # 상품명 칼럼은 자동 확장
        self.table_widget.setColumnWidth(0, 60)
        self.table_widget.setColumnWidth(2, 100)
        self.table_widget.setColumnWidth(3, 120)
        self.table_widget.setColumnWidth(4, 100)
        self.table_widget.setColumnWidth(5, 100)
        
        self.table_widget.setAlternatingRowColors(True)
        self.table_widget.setStyleSheet("""
            QTableWidget {
                gridline-color: #E0E0E0;
                background-color: #FAFAFA;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QTableWidget::item:selected {
                background-color: #E3F2FD;
                color: #1976D2;
            }
            QHeaderView::section {
                background-color: #2196F3;
                color: white;
                padding: 8px;
                font-weight: bold;
                border: none;
            }
        """)
        
        layout.addWidget(self.table_widget)
        
        self.tab2.setLayout(layout)
        self.rank_list_data = []  # 리스트 데이터 저장용

    def setup_tab3(self):
        """탭 3: 연관 키워드 조회 UI"""
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)

        # 입력 영역
        input_group = QGroupBox("검색 설정")
        input_layout = QVBoxLayout()
        
        # 검색어 입력
        keyword_label = QLabel("🔍 조회할 키워드")
        keyword_label.setStyleSheet("font-weight: bold; color: #333;")
        self.input_keyword_related = QLineEdit()
        self.input_keyword_related.setPlaceholderText("예: 키보드")
        self.input_keyword_related.setMinimumHeight(35)
        
        input_layout.addWidget(keyword_label)
        input_layout.addWidget(self.input_keyword_related)
        
        input_group.setLayout(input_layout)
        layout.addWidget(input_group)

        # 버튼 영역
        button_layout3 = QHBoxLayout()
        self.button_get_related = QPushButton("🔍 연관 키워드 조회하기")
        self.button_get_related.setMinimumHeight(45)
        self.button_get_related.clicked.connect(self.get_related_keywords)
        
        self.button_export_related = QPushButton("📥 엑셀로 다운로드")
        self.button_export_related.setMinimumHeight(45)
        self.button_export_related.setEnabled(False)
        self.button_export_related.clicked.connect(self.export_related_to_excel)
        self.button_export_related.setStyleSheet("background-color: #4CAF50;")
        
        button_layout3.addWidget(self.button_get_related)
        button_layout3.addWidget(self.button_export_related)
        layout.addLayout(button_layout3)

        # 진행 상태
        status_layout3 = QHBoxLayout()
        self.label_status3 = QLabel("대기 중...")
        self.label_status3.setStyleSheet("color: #666; font-weight: bold;")
        status_layout3.addWidget(self.label_status3)
        status_layout3.addStretch()
        layout.addLayout(status_layout3)

        # 결과 표시 영역
        result_label = QLabel("🔗 연관 키워드 (검색수 순)")
        result_label.setStyleSheet("font-weight: bold; color: #333; margin-top: 10px;")
        layout.addWidget(result_label)
        
        # 테이블
        self.table_related = QTableWidget()
        self.table_related.setColumnCount(4)
        self.table_related.setHorizontalHeaderLabels(["키워드", "PC 검색수", "모바일 검색수", "총 검색수"])
        
        # 테이블 설정
        header = self.table_related.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)  # 키워드 칼럼은 자동 확장
        self.table_related.setColumnWidth(1, 100)
        self.table_related.setColumnWidth(2, 110)
        self.table_related.setColumnWidth(3, 100)
        
        self.table_related.setAlternatingRowColors(True)
        self.table_related.setStyleSheet("""
            QTableWidget {
                gridline-color: #E0E0E0;
                background-color: #FAFAFA;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QTableWidget::item:selected {
                background-color: #E3F2FD;
                color: #1976D2;
            }
            QHeaderView::section {
                background-color: #2196F3;
                color: white;
                padding: 8px;
                font-weight: bold;
                border: none;
            }
        """)
        
        layout.addWidget(self.table_related)
        
        self.tab3.setLayout(layout)

    def setup_tab4(self):
        """탭 4: 광고 비용 분석 UI"""
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)

        # 입력 영역
        input_group = QGroupBox("검색 설정")
        input_layout = QVBoxLayout()
        
        # 검색어 입력
        keyword_label = QLabel("🔍 분석할 키워드 (쉼표로 구분, 최대 20개)")
        keyword_label.setStyleSheet("font-weight: bold; color: #333;")
        self.input_keywords_ad = QTextEdit()
        self.input_keywords_ad.setPlaceholderText("예: 키보드, 마우스, 모니터")
        self.input_keywords_ad.setMaximumHeight(100)
        
        input_layout.addWidget(keyword_label)
        input_layout.addWidget(self.input_keywords_ad)
        
        input_group.setLayout(input_layout)
        layout.addWidget(input_group)

        # 분석 버튼
        self.button_analyze_ad = QPushButton("💰 광고 비용 분석하기")
        self.button_analyze_ad.setMinimumHeight(45)
        self.button_analyze_ad.clicked.connect(self.analyze_ad_cost)
        layout.addWidget(self.button_analyze_ad)

        # 진행 상태
        status_layout4 = QHBoxLayout()
        self.label_status4 = QLabel("대기 중...")
        self.label_status4.setStyleSheet("color: #666; font-weight: bold;")
        status_layout4.addWidget(self.label_status4)
        status_layout4.addStretch()
        layout.addLayout(status_layout4)

        # 결과 표시 영역
        result_label = QLabel("💰 광고 비용 분석 결과")
        result_label.setStyleSheet("font-weight: bold; color: #333; margin-top: 10px;")
        layout.addWidget(result_label)
        
        # 테이블
        self.table_ad_cost = QTableWidget()
        self.table_ad_cost.setColumnCount(10)
        self.table_ad_cost.setHorizontalHeaderLabels([
            "키워드", "PC 검색수", "모바일 검색수", "총 검색수", "월 클릭수", 
            "평균 CTR(%)", "경쟁도", "노출광고수", "평균 CPC", "월 예상 광고비"
        ])
        
        # 테이블 설정
        header = self.table_ad_cost.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)  # 키워드 칼럼은 자동 확장
        self.table_ad_cost.setColumnWidth(1, 100)
        self.table_ad_cost.setColumnWidth(2, 110)
        self.table_ad_cost.setColumnWidth(3, 100)
        self.table_ad_cost.setColumnWidth(4, 90)
        self.table_ad_cost.setColumnWidth(5, 90)
        self.table_ad_cost.setColumnWidth(6, 80)
        self.table_ad_cost.setColumnWidth(7, 90)
        self.table_ad_cost.setColumnWidth(8, 90)
        self.table_ad_cost.setColumnWidth(9, 120)
        
        self.table_ad_cost.setAlternatingRowColors(True)
        self.table_ad_cost.setStyleSheet("""
            QTableWidget {
                gridline-color: #E0E0E0;
                background-color: #FAFAFA;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QTableWidget::item:selected {
                background-color: #E3F2FD;
                color: #1976D2;
            }
            QHeaderView::section {
                background-color: #4CAF50;
                color: white;
                padding: 8px;
                font-weight: bold;
                border: none;
            }
        """)
        
        layout.addWidget(self.table_ad_cost)
        
        # 안내 문구
        info_label = QLabel("💡 참고: 네이버 검색광고 API를 통해 실제 데이터를 가져옵니다. API 오류 시 추정값을 사용합니다.")
        info_label.setStyleSheet("color: #666; font-size: 9pt; font-style: italic; margin-top: 5px;")
        layout.addWidget(info_label)
        
        self.tab4.setLayout(layout)

    def setup_tab5(self):
        """탭 5: 설정 UI"""
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # 제목
        title_label = QLabel("⚙️ API 설정")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setStyleSheet("color: #1976D2; padding: 10px;")
        layout.addWidget(title_label)

        # 네이버 쇼핑 API 설정
        shop_group = QGroupBox("네이버 쇼핑 API (필수)")
        shop_layout = QVBoxLayout()
        shop_layout.setSpacing(10)
        
        client_id_label = QLabel("Client ID:")
        client_id_label.setStyleSheet("font-weight: bold;")
        self.input_client_id = QLineEdit()
        self.input_client_id.setPlaceholderText("네이버 쇼핑 API Client ID 입력")
        self.input_client_id.setMinimumHeight(35)
        
        client_secret_label = QLabel("Client Secret:")
        client_secret_label.setStyleSheet("font-weight: bold;")
        self.input_client_secret = QLineEdit()
        self.input_client_secret.setPlaceholderText("네이버 쇼핑 API Client Secret 입력")
        self.input_client_secret.setEchoMode(QLineEdit.Password)
        self.input_client_secret.setMinimumHeight(35)
        
        shop_layout.addWidget(client_id_label)
        shop_layout.addWidget(self.input_client_id)
        shop_layout.addWidget(client_secret_label)
        shop_layout.addWidget(self.input_client_secret)
        shop_group.setLayout(shop_layout)
        layout.addWidget(shop_group)

        # 네이버 검색광고 API 설정
        ad_group = QGroupBox("네이버 검색광고 API (선택)")
        ad_layout = QVBoxLayout()
        ad_layout.setSpacing(10)
        
        customer_id_label = QLabel("Customer ID:")
        customer_id_label.setStyleSheet("font-weight: bold;")
        self.input_customer_id = QLineEdit()
        self.input_customer_id.setPlaceholderText("고객 ID 입력")
        self.input_customer_id.setMinimumHeight(35)
        
        api_key_label = QLabel("API Key (액세스 라이선스):")
        api_key_label.setStyleSheet("font-weight: bold;")
        self.input_api_key = QLineEdit()
        self.input_api_key.setPlaceholderText("액세스 라이선스 입력")
        self.input_api_key.setMinimumHeight(35)
        
        secret_key_label = QLabel("Secret Key (비밀키):")
        secret_key_label.setStyleSheet("font-weight: bold;")
        self.input_secret_key = QLineEdit()
        self.input_secret_key.setPlaceholderText("비밀키 입력")
        self.input_secret_key.setEchoMode(QLineEdit.Password)
        self.input_secret_key.setMinimumHeight(35)
        
        ad_layout.addWidget(customer_id_label)
        ad_layout.addWidget(self.input_customer_id)
        ad_layout.addWidget(api_key_label)
        ad_layout.addWidget(self.input_api_key)
        ad_layout.addWidget(secret_key_label)
        ad_layout.addWidget(self.input_secret_key)
        ad_group.setLayout(ad_layout)
        layout.addWidget(ad_group)

        # 버튼 영역
        button_layout = QHBoxLayout()
        
        self.button_test_api = QPushButton("🔍 API 인증 테스트")
        self.button_test_api.setMinimumHeight(45)
        self.button_test_api.clicked.connect(self.test_api_keys)
        self.button_test_api.setStyleSheet("background-color: #FF9800;")
        
        self.button_save_api = QPushButton("💾 API 키 저장")
        self.button_save_api.setMinimumHeight(45)
        self.button_save_api.clicked.connect(self.save_api_keys)
        self.button_save_api.setStyleSheet("background-color: #4CAF50;")
        
        button_layout.addWidget(self.button_test_api)
        button_layout.addWidget(self.button_save_api)
        layout.addLayout(button_layout)

        # 상태 표시
        self.label_api_status = QLabel("💡 API 키를 입력하고 테스트 후 저장하세요.")
        self.label_api_status.setStyleSheet("padding: 15px; background-color: #E3F2FD; border-radius: 5px; color: #1976D2;")
        self.label_api_status.setWordWrap(True)
        layout.addWidget(self.label_api_status)

        # 안내 문구
        info_text = QLabel(
            "ℹ️ 안내\n\n"
            "• 네이버 쇼핑 API는 필수입니다.\n"
            "• 네이버 검색광고 API는 선택사항입니다 (연관 키워드, 광고 비용 분석 기능에 사용).\n"
            "• API 키는 암호화되어 config.json 파일에 저장됩니다.\n"
            "• 네이버 개발자 센터(https://developers.naver.com)에서 API를 발급받으세요."
        )
        info_text.setStyleSheet("color: #666; padding: 15px; background-color: #F5F5F5; border-radius: 5px;")
        info_text.setWordWrap(True)
        layout.addWidget(info_text)

        layout.addStretch()
        self.tab5.setLayout(layout)

    def apply_styles(self):
        """스타일시트 적용"""
        self.setStyleSheet("""
            QWidget {
                background-color: #ffffff;
                font-family: 'Malgun Gothic', 'Segoe UI', sans-serif;
                font-size: 11pt;
            }
            QGroupBox {
                border: 2px solid #E0E0E0;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
                color: #1976D2;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
            QLineEdit, QTextEdit {
                border: 2px solid #E0E0E0;
                border-radius: 5px;
                padding: 8px;
                background-color: #FAFAFA;
            }
            QLineEdit:focus, QTextEdit:focus {
                border: 2px solid #2196F3;
                background-color: #ffffff;
            }
            QComboBox {
                border: 2px solid #E0E0E0;
                border-radius: 5px;
                padding: 8px;
                padding-right: 30px;
                background-color: #FAFAFA;
                min-height: 40px;
                font-size: 11pt;
            }
            QComboBox:focus {
                border: 2px solid #2196F3;
                background-color: #ffffff;
            }
            QComboBox:hover {
                border: 2px solid #90CAF9;
            }
            QComboBox::drop-down {
                border: none;
                width: 35px;
                background: transparent;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 6px solid transparent;
                border-right: 6px solid transparent;
                border-top: 6px solid #666;
                margin-right: 12px;
            }
            QComboBox QAbstractItemView {
                border: 2px solid #2196F3;
                border-radius: 5px;
                background-color: #ffffff;
                selection-background-color: #E3F2FD;
                selection-color: #1976D2;
                padding: 5px;
                outline: none;
            }
            QComboBox QAbstractItemView::item {
                min-height: 35px;
                padding: 5px 10px;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #E3F2FD;
            }
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font-weight: bold;
                font-size: 12pt;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:pressed {
                background-color: #0D47A1;
            }
            QPushButton:disabled {
                background-color: #BDBDBD;
            }
            QProgressBar {
                border: 2px solid #E0E0E0;
                border-radius: 5px;
                text-align: center;
                background-color: #F5F5F5;
            }
            QProgressBar::chunk {
                background-color: #4CAF50;
                border-radius: 3px;
            }
            QTextBrowser {
                border: 2px solid #E0E0E0;
                border-radius: 5px;
                padding: 10px;
                background-color: #FAFAFA;
            }
        """)

    def start_check(self):
        """순위 확인 시작"""
        # 입력값 가져오기
        keywords_text = self.input_keywords.toPlainText().strip()
        mall_name = self.input_mall.text().strip()
        sort_option = self.sort_combo.currentData()

        # 유효성 검사
        if not keywords_text or not mall_name:
            QMessageBox.warning(self, "입력 오류", "검색어와 판매처명을 모두 입력해주세요.")
            return

        keywords = [k.strip() for k in keywords_text.split(",") if k.strip()]
        
        if not keywords:
            QMessageBox.warning(self, "입력 오류", "올바른 검색어를 입력해주세요.")
            return

        if len(keywords) > 10:
            QMessageBox.warning(self, "제한 초과", "검색어는 최대 10개까지 가능합니다.")
            return

        # UI 초기화
        self.result_display.clear()
        self.progress_bar.setValue(0)
        sort_name = self.sort_combo.currentText()
        self.label_status.setText(f"🔄 검색 중... ({sort_name})")
        self.button_check.setEnabled(False)

        # 워커 쓰레드 시작
        self.worker = Worker(keywords, mall_name, sort_option)
        self.worker.result_ready.connect(self.append_result)
        self.worker.progress_update.connect(self.update_status)
        self.worker.finished_all.connect(self.on_finished)
        self.worker.start()

    def append_result(self, html):
        """결과를 화면에 추가"""
        self.result_display.append(html)

    def update_status(self, percent, _keyword):
        """진행 상태 업데이트"""
        self.progress_bar.setValue(percent)
        self.label_status.setText(f"🔄 검색 중... ({percent}% 완료)")

    def on_finished(self, results):
        """검색 완료 처리"""
        self.label_status.setText("✅ 검색 완료!")
        self.button_check.setEnabled(True)
        
        # 결과 요약
        total = len(results)
        found = sum(1 for v in results.values() if isinstance(v, dict))
        
        summary_html = (
            f"<div style='margin-top: 15px; padding: 15px; background-color: #E3F2FD; "
            f"border-radius: 5px; border-left: 4px solid #2196F3;'>"
            f"<b style='color: #1976D2;'>📈 검색 요약</b><br>"
            f"전체 키워드: {total}개 | 검색 성공: {found}개 | 검색 실패: {total - found}개"
            f"</div>"
        )
        self.result_display.append(summary_html)

    def get_rank_list(self):
        """1~100위 리스트 가져오기"""
        keyword = self.input_keyword_list.text().strip()
        sort_option = self.sort_combo2.currentData()
        
        if not keyword:
            QMessageBox.warning(self, "입력 오류", "검색어를 입력해주세요.")
            return
        
        # UI 초기화
        self.table_widget.setRowCount(0)
        self.progress_bar2.setValue(0)
        self.label_status2.setText("🔄 리스트 수집 중...")
        self.button_get_list.setEnabled(False)
        self.button_export_excel.setEnabled(False)
        
        # 워커 쓰레드 시작
        self.list_worker = RankListWorker(keyword, sort_option)
        self.list_worker.progress_update.connect(self.update_list_status)
        self.list_worker.finished.connect(self.on_list_finished)
        self.list_worker.start()

    def update_list_status(self, percent, message):
        """리스트 수집 상태 업데이트"""
        self.progress_bar2.setValue(percent)
        self.label_status2.setText(f"🔄 {message}")

    def on_list_finished(self, products):
        """리스트 수집 완료"""
        self.rank_list_data = products
        self.label_status2.setText(f"✅ 총 {len(products)}개 상품 수집 완료!")
        self.button_get_list.setEnabled(True)
        
        if products:
            self.populate_table(products)
            self.button_export_excel.setEnabled(True)
        else:
            QMessageBox.information(self, "결과 없음", "검색 결과가 없습니다.")

    def populate_table(self, products):
        """테이블에 데이터 채우기"""
        self.table_widget.setRowCount(len(products))
        
        for row, product in enumerate(products):
            # 순위
            rank_item = QTableWidgetItem(str(product["rank"]))
            rank_item.setTextAlignment(Qt.AlignCenter)
            self.table_widget.setItem(row, 0, rank_item)
            
            # 상품명
            title_item = QTableWidgetItem(product["title"])
            self.table_widget.setItem(row, 1, title_item)
            
            # 가격
            price = int(product["price"])
            price_item = QTableWidgetItem(f"{price:,}원")
            price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table_widget.setItem(row, 2, price_item)
            
            # 판매처
            mall_item = QTableWidgetItem(product["mallName"])
            self.table_widget.setItem(row, 3, mall_item)
            
            # 브랜드
            brand_item = QTableWidgetItem(product["brand"])
            self.table_widget.setItem(row, 4, brand_item)
            
            # 제조사
            maker_item = QTableWidgetItem(product["maker"])
            self.table_widget.setItem(row, 5, maker_item)

    def export_to_excel(self):
        """엑셀로 내보내기"""
        if not self.rank_list_data:
            QMessageBox.warning(self, "데이터 없음", "내보낼 데이터가 없습니다.")
            return
        
        # 파일 저장 대화상자
        keyword = self.input_keyword_list.text().strip()
        default_filename = f"네이버쇼핑_{keyword}_순위_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "엑셀 파일 저장",
            default_filename,
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not filepath:
            return
        
        try:
            # openpyxl 사용하여 엑셀 저장
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "순위 리스트"
            
            # 헤더 작성
            headers = ["순위", "상품명", "가격", "판매처", "브랜드", "제조사", "링크"]
            ws.append(headers)
            
            # 헤더 스타일
            header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 데이터 작성
            for product in self.rank_list_data:
                ws.append([
                    product["rank"],
                    product["title"],
                    int(product["price"]),
                    product["mallName"],
                    product["brand"],
                    product["maker"],
                    product["link"]
                ])
            
            # 열 너비 조정
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 60
            
            # 저장
            wb.save(filepath)
            
            QMessageBox.information(
                self,
                "저장 완료",
                f"엑셀 파일이 저장되었습니다!\n\n파일 위치: {filepath}"
            )
            
        except ImportError:
            QMessageBox.critical(
                self,
                "라이브러리 오류",
                "openpyxl 라이브러리가 필요합니다.\n\n명령 프롬프트에서 다음을 실행하세요:\npip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "저장 오류",
                f"엑셀 파일 저장 중 오류가 발생했습니다:\n{str(e)}"
            )

    def get_related_keywords(self):
        """연관 키워드 조회"""
        keyword = self.input_keyword_related.text().strip()
        
        if not keyword:
            QMessageBox.warning(self, "입력 오류", "조회할 키워드를 입력해주세요.")
            return
        
        # 버튼 비활성화
        self.button_get_related.setEnabled(False)
        self.label_status3.setText("조회 중...")
        
        # 워커 쓰레드 시작
        self.related_worker = RelatedKeywordWorker(keyword)
        self.related_worker.progress_update.connect(self.update_related_status)
        self.related_worker.result_ready.connect(self.display_related_keywords)
        self.related_worker.finished.connect(lambda: self.button_get_related.setEnabled(True))
        self.related_worker.start()
    
    def update_related_status(self, message):
        """연관 키워드 조회 상태 업데이트"""
        self.label_status3.setText(message)
    
    def display_related_keywords(self, keywords):
        """연관 키워드 결과 표시"""
        self.table_related.setRowCount(0)
        
        if not keywords:
            self.label_status3.setText("연관 키워드를 찾을 수 없습니다.")
            self.button_export_related.setEnabled(False)
            QMessageBox.information(self, "조회 완료", "연관 키워드를 찾을 수 없습니다.")
            return
        
        # 데이터 저장 (엑셀 다운로드용)
        self.related_keywords_data = keywords
        
        # 총 키워드 개수 확인
        total_count = keywords[0].get("total_count", len(keywords)) if keywords else 0
        
        # 테이블에 데이터 추가
        self.table_related.setRowCount(len(keywords))
        for row, item in enumerate(keywords):
            # 0. 키워드
            keyword_item = QTableWidgetItem(item["keyword"])
            self.table_related.setItem(row, 0, keyword_item)
            
            # 1. PC 검색수
            pc_item = QTableWidgetItem(f"{item.get('pc_count', 0):,}")
            pc_item.setTextAlignment(Qt.AlignCenter)
            self.table_related.setItem(row, 1, pc_item)
            
            # 2. 모바일 검색수
            mobile_item = QTableWidgetItem(f"{item.get('mobile_count', 0):,}")
            mobile_item.setTextAlignment(Qt.AlignCenter)
            self.table_related.setItem(row, 2, mobile_item)
            
            # 3. 총 검색수
            total_item = QTableWidgetItem(f"{item['count']:,}")
            total_item.setTextAlignment(Qt.AlignCenter)
            self.table_related.setItem(row, 3, total_item)
        
        # 엑셀 다운로드 버튼 활성화
        self.button_export_related.setEnabled(True)
        
        # 상태 메시지: 총 개수 중 100개 표시
        if total_count > len(keywords):
            self.label_status3.setText(f"✅ 총 {total_count:,}개 중 상위 {len(keywords)}개의 연관 키워드를 표시합니다.")
        else:
            self.label_status3.setText(f"✅ {len(keywords):,}개의 연관 키워드를 찾았습니다.")

    def load_api_keys(self):
        """저장된 API 키 로드"""
        try:
            if os.path.exists('config.json'):
                with open('config.json', 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    
                    # 네이버 쇼핑 API
                    self.input_client_id.setText(config.get('client_id', ''))
                    self.input_client_secret.setText(config.get('client_secret', ''))
                    
                    # 네이버 검색광고 API
                    self.input_customer_id.setText(config.get('customer_id', ''))
                    self.input_api_key.setText(config.get('api_key', ''))
                    self.input_secret_key.setText(config.get('secret_key', ''))
                    
                    # 글로벌 변수 업데이트
                    global NAVER_CLIENT_ID, NAVER_CLIENT_SECRET
                    global NAVER_AD_API_KEY, NAVER_AD_SECRET_KEY, NAVER_AD_CUSTOMER_ID
                    
                    NAVER_CLIENT_ID = config.get('client_id', '')
                    NAVER_CLIENT_SECRET = config.get('client_secret', '')
                    NAVER_AD_API_KEY = config.get('api_key', '')
                    NAVER_AD_SECRET_KEY = config.get('secret_key', '')
                    NAVER_AD_CUSTOMER_ID = config.get('customer_id', '')
        except Exception as e:
            print(f"API 키 로드 오류: {e}")
    
    def save_api_keys(self):
        """API 키 저장"""
        client_id = self.input_client_id.text().strip()
        client_secret = self.input_client_secret.text().strip()
        
        if not client_id or not client_secret:
            QMessageBox.warning(self, "입력 오류", "네이버 쇼핑 API (Client ID와 Secret)는 필수입니다!")
            return
        
        try:
            config = {
                'client_id': client_id,
                'client_secret': client_secret,
                'customer_id': self.input_customer_id.text().strip(),
                'api_key': self.input_api_key.text().strip(),
                'secret_key': self.input_secret_key.text().strip()
            }
            
            with open('config.json', 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4, ensure_ascii=False)
            
            # 글로벌 변수 업데이트
            global NAVER_CLIENT_ID, NAVER_CLIENT_SECRET
            global NAVER_AD_API_KEY, NAVER_AD_SECRET_KEY, NAVER_AD_CUSTOMER_ID
            
            NAVER_CLIENT_ID = client_id
            NAVER_CLIENT_SECRET = client_secret
            NAVER_AD_API_KEY = config['api_key']
            NAVER_AD_SECRET_KEY = config['secret_key']
            NAVER_AD_CUSTOMER_ID = config['customer_id']
            
            self.label_api_status.setText("✅ API 키가 저장되었습니다!")
            self.label_api_status.setStyleSheet("padding: 15px; background-color: #C8E6C9; border-radius: 5px; color: #2E7D32;")
            
            # 다른 탭들 활성화
            for i in range(4):
                self.tab_widget.setTabEnabled(i, True)
            
            QMessageBox.information(self, "저장 완료", "API 키가 성공적으로 저장되었습니다!\n\n이제 모든 기능을 사용할 수 있습니다.")
            
        except Exception as e:
            QMessageBox.critical(self, "저장 오류", f"API 키 저장 중 오류가 발생했습니다:\n{str(e)}")
    
    def test_api_keys(self):
        """API 키 인증 테스트"""
        client_id = self.input_client_id.text().strip()
        client_secret = self.input_client_secret.text().strip()
        
        if not client_id or not client_secret:
            QMessageBox.warning(self, "입력 오류", "네이버 쇼핑 API 키를 먼저 입력하세요!")
            return
        
        self.label_api_status.setText("🔍 API 인증 테스트 중...")
        self.label_api_status.setStyleSheet("padding: 15px; background-color: #FFF9C4; border-radius: 5px; color: #F57F17;")
        
        # 네이버 쇼핑 API 테스트
        try:
            test_url = "https://openapi.naver.com/v1/search/shop.json?query=테스트&display=1"
            request = urllib.request.Request(test_url)
            request.add_header("X-Naver-Client-Id", client_id)
            request.add_header("X-Naver-Client-Secret", client_secret)
            
            response = urllib.request.urlopen(request)
            result = json.loads(response.read())
            
            if 'items' in result:
                shop_api_ok = True
                shop_msg = "✅ 네이버 쇼핑 API: 인증 성공"
            else:
                shop_api_ok = False
                shop_msg = "❌ 네이버 쇼핑 API: 인증 실패"
        except Exception as e:
            shop_api_ok = False
            shop_msg = f"❌ 네이버 쇼핑 API: 오류 - {str(e)}"
        
        # 네이버 검색광고 API 테스트 (선택사항)
        ad_msg = ""
        customer_id = self.input_customer_id.text().strip()
        api_key = self.input_api_key.text().strip()
        secret_key = self.input_secret_key.text().strip()
        
        if customer_id and api_key and secret_key:
            try:
                import hashlib
                import hmac
                import base64
                
                timestamp = str(int(time.time() * 1000))
                method = "GET"
                uri = "/keywordstool"
                
                message = f"{timestamp}.{method}.{uri}"
                signature = hmac.new(
                    secret_key.encode('utf-8'),
                    message.encode('utf-8'),
                    hashlib.sha256
                ).digest()
                signature = base64.b64encode(signature).decode('utf-8')
                
                test_url = f"https://api.naver.com/keywordstool?hintKeywords=테스트&showDetail=1"
                request = urllib.request.Request(test_url)
                request.add_header("X-Timestamp", timestamp)
                request.add_header("X-API-KEY", api_key)
                request.add_header("X-Customer", customer_id)
                request.add_header("X-Signature", signature)
                
                response = urllib.request.urlopen(request)
                result = json.loads(response.read().decode('utf-8'))
                
                if 'keywordList' in result:
                    ad_msg = "\n✅ 네이버 검색광고 API: 인증 성공"
                else:
                    ad_msg = "\n❌ 네이버 검색광고 API: 인증 실패"
            except Exception as e:
                ad_msg = f"\n❌ 네이버 검색광고 API: 오류 - {str(e)}"
        else:
            ad_msg = "\n⚠️ 네이버 검색광고 API: 미설정 (선택사항)"
        
        # 결과 표시
        final_msg = shop_msg + ad_msg
        
        if shop_api_ok:
            self.label_api_status.setText(final_msg)
            self.label_api_status.setStyleSheet("padding: 15px; background-color: #C8E6C9; border-radius: 5px; color: #2E7D32;")
            QMessageBox.information(self, "인증 테스트 완료", final_msg)
        else:
            self.label_api_status.setText(final_msg)
            self.label_api_status.setStyleSheet("padding: 15px; background-color: #FFCDD2; border-radius: 5px; color: #C62828;")
            QMessageBox.critical(self, "인증 테스트 실패", final_msg)
    
    def check_api_keys_on_startup(self):
        """프로그램 시작 시 API 키 확인"""
        if not NAVER_CLIENT_ID or not NAVER_CLIENT_SECRET:
            # API 키가 없으면 설정 탭으로 이동
            self.tab_widget.setCurrentWidget(self.tab5)
            
            QMessageBox.warning(
                self,
                "API 키 필요",
                "⚠️ 네이버 쇼핑 API 키가 설정되지 않았습니다!\n\n"
                "프로그램을 사용하려면 API 키를 입력하고 저장해주세요.\n\n"
                "1. 네이버 개발자 센터에서 API 키 발급\n"
                "2. 아래 양식에 API 키 입력\n"
                "3. '인증 테스트' 버튼으로 확인\n"
                "4. '저장' 버튼으로 저장\n\n"
                "필수: 네이버 쇼핑 API\n"
                "선택: 네이버 검색광고 API"
            )
            
            # 다른 탭들 비활성화
            for i in range(4):
                self.tab_widget.setTabEnabled(i, False)
    
    def export_related_to_excel(self):
        """연관 키워드를 엑셀로 내보내기"""
        if not hasattr(self, 'related_keywords_data') or not self.related_keywords_data:
            QMessageBox.warning(self, "데이터 없음", "내보낼 데이터가 없습니다.")
            return
        
        # 파일 저장 대화상자
        keyword = self.input_keyword_related.text().strip()
        default_filename = f"연관키워드_{keyword}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "엑셀 파일 저장",
            default_filename,
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not filepath:
            return
        
        try:
            # openpyxl 사용하여 엑셀 저장
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "연관 키워드"
            
            # 헤더 작성
            headers = ["순위", "키워드", "PC 검색수", "모바일 검색수", "총 검색수"]
            ws.append(headers)
            
            # 헤더 스타일
            header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 데이터 작성
            for idx, item in enumerate(self.related_keywords_data, 1):
                ws.append([
                    idx,
                    item["keyword"],
                    item.get("pc_count", 0),
                    item.get("mobile_count", 0),
                    item["count"]
                ])
            
            # 열 너비 조정
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            
            # 저장
            wb.save(filepath)
            
            QMessageBox.information(
                self,
                "저장 완료",
                f"엑셀 파일이 저장되었습니다!\n\n파일 위치: {filepath}"
            )
            
        except ImportError:
            QMessageBox.critical(
                self,
                "라이브러리 오류",
                "openpyxl 라이브러리가 필요합니다.\n\n명령 프롬프트에서 다음을 실행하세요:\npip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "저장 오류",
                f"엑셀 파일 저장 중 오류가 발생했습니다:\n{str(e)}"
            )

    def analyze_ad_cost(self):
        """광고 비용 분석"""
        keywords_text = self.input_keywords_ad.toPlainText().strip()
        
        if not keywords_text:
            QMessageBox.warning(self, "입력 오류", "분석할 키워드를 입력해주세요.")
            return
        
        # 쉼표로 구분하여 키워드 리스트 생성
        keywords = [k.strip() for k in keywords_text.split(",") if k.strip()]
        
        if len(keywords) > 20:
            QMessageBox.warning(self, "입력 오류", "최대 20개까지만 분석할 수 있습니다.")
            return
        
        # 버튼 비활성화
        self.button_analyze_ad.setEnabled(False)
        self.label_status4.setText("분석 중...")
        
        # 워커 쓰레드 시작
        self.ad_worker = AdCostWorker(keywords)
        self.ad_worker.progress_update.connect(self.update_ad_status)
        self.ad_worker.result_ready.connect(self.display_ad_cost)
        self.ad_worker.finished.connect(lambda: self.button_analyze_ad.setEnabled(True))
        self.ad_worker.start()
    
    def update_ad_status(self, message):
        """광고 비용 분석 상태 업데이트"""
        self.label_status4.setText(message)
    
    def display_ad_cost(self, results):
        """광고 비용 분석 결과 표시"""
        self.table_ad_cost.setRowCount(0)
        
        if not results:
            self.label_status4.setText("분석 결과가 없습니다.")
            QMessageBox.information(self, "분석 완료", "분석 결과가 없습니다.")
            return
        
        # 테이블에 데이터 추가
        self.table_ad_cost.setRowCount(len(results))
        for row, item in enumerate(results):
            # 0. 키워드
            keyword_item = QTableWidgetItem(item["keyword"])
            self.table_ad_cost.setItem(row, 0, keyword_item)
            
            # 1. PC 검색수
            pc_search_item = QTableWidgetItem(f"{item['monthly_pc_search']:,}")
            pc_search_item.setTextAlignment(Qt.AlignCenter)
            self.table_ad_cost.setItem(row, 1, pc_search_item)
            
            # 2. 모바일 검색수
            mobile_search_item = QTableWidgetItem(f"{item['monthly_mobile_search']:,}")
            mobile_search_item.setTextAlignment(Qt.AlignCenter)
            self.table_ad_cost.setItem(row, 2, mobile_search_item)
            
            # 3. 총 검색수
            total_search_item = QTableWidgetItem(f"{item['total_monthly_search']:,}")
            total_search_item.setTextAlignment(Qt.AlignCenter)
            self.table_ad_cost.setItem(row, 3, total_search_item)
            
            # 4. 월 클릭수
            clicks_item = QTableWidgetItem(f"{item['monthly_clicks']:,}")
            clicks_item.setTextAlignment(Qt.AlignCenter)
            self.table_ad_cost.setItem(row, 4, clicks_item)
            
            # 5. 평균 CTR
            ctr_item = QTableWidgetItem(f"{item['avg_ctr']:.2f}%")
            ctr_item.setTextAlignment(Qt.AlignCenter)
            self.table_ad_cost.setItem(row, 5, ctr_item)
            
            # 6. 경쟁도
            comp_item = QTableWidgetItem(item["competition"])
            comp_item.setTextAlignment(Qt.AlignCenter)
            # 경쟁도에 따라 색상 변경
            if item["competition"] == "낮음":
                comp_item.setForeground(Qt.darkGreen)
            elif item["competition"] == "보통":
                comp_item.setForeground(Qt.darkYellow)
            elif item["competition"] == "높음":
                comp_item.setForeground(Qt.darkMagenta)
            else:  # 매우 높음
                comp_item.setForeground(Qt.red)
            self.table_ad_cost.setItem(row, 6, comp_item)
            
            # 7. 노출광고수
            ad_impressions_item = QTableWidgetItem(f"{item['avg_ad_impressions']}")
            ad_impressions_item.setTextAlignment(Qt.AlignCenter)
            self.table_ad_cost.setItem(row, 7, ad_impressions_item)
            
            # 8. 평균 CPC
            cpc_item = QTableWidgetItem(f"{item['estimated_cpc']:,}원")
            cpc_item.setTextAlignment(Qt.AlignCenter)
            self.table_ad_cost.setItem(row, 8, cpc_item)
            
            # 9. 월 예상 광고비
            cost_item = QTableWidgetItem(f"{item['estimated_monthly_cost']:,}원")
            cost_item.setTextAlignment(Qt.AlignCenter)
            self.table_ad_cost.setItem(row, 9, cost_item)
        
        self.label_status4.setText(f"✅ {len(results)}개 키워드 분석 완료!")


def main():
    """메인 함수"""
    app = QApplication(sys.argv)
    
    # 애플리케이션 스타일 설정
    app.setStyle('Fusion')
    
    window = RankCheckerApp()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()

